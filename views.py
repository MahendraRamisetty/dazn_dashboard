from django.shortcuts import render, redirect
from .forms import UploadFileForm
from .models import UploadedFile
import pandas as pd
import urllib.parse
from openpyxl import load_workbook
from django.core.paginator import Paginator
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for Matplotlib
import matplotlib.pyplot as plt
import io
import base64
import pandas as pd
import matplotlib.pyplot as plt
import io
import base64
import numpy as np
from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import base64
from .summary import (
    aggregate_matchday,
    get_social_media_platform_data,
    create_social_media_platform_bar_chart,
    create_matchday_line_plot,
    get_monthly_totals,
    create_monthly_totals_line_plot,
    create_bar_chart,
    generate_top_10_piracy_sources_graph,
    generate_top_10_search_treemap,
    generate_top_10_domains_bar_chart
)
from dashboardData.internet import generate_internet_graphs
from .mobile import Mobile_properties_data, MobileApp_bar_chart, mobile_monthly_totals_line_plot , mobile_monthly_totals, Mobile_top_apps_by_downloads, create_mobile_top_apps_bar_chart, mobile_aggregate_matchday, mobile_create_matchday_line_plot

# ✅ Import all required functions from telegram.py
from .telegram import (
    calculate_telegram_summary, 
    get_top_telegram_properties, 
    telegram_monthly_totals, 
    create_telegram_top_fixtures_bar_chart, 
    create_telegram_channel_type_pie_chart,
    get_telegram_platform_data, 
    get_telegram_top_fixtures, 
    telegram_domains_by_subscribers, 
    create_treemap_chart_telegram, 
    create_enhanced_matchday_line_plot, 
    aggregate_matchday_data, 
    telegram_monthly_totals_line_plot,
    top_fixtures_donut_chart,
    top_fixtures_graph_donut_chart,
    create_channel_type_pie_chart,
    create_top_property_bar_chart,
)

from django.urls import reverse
import urllib.parse
from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import base64
import os
import json
import logging
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import requests
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PIL import Image
import logging
import requests
import base64
from datetime import datetime
from io import BytesIO
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

LOGIC_APP_URL = "https://prod-43.northeurope.logic.azure.com:443/workflows/b9ccb16083f04f6bb07a9679303a7ffd/triggers/HTTPReceived/paths/invoke?api-version=2016-10-01&sp=%2Ftriggers%2FHTTPReceived%2Frun&sv=1.0&sig=VVYJDahy4jEeF4NpFqBuZgTHh1cFZ5L0lAcBIv6MuWY"

@csrf_exempt
def send_pdf_email(request):
    if request.method != "POST":
        return JsonResponse({'error': 'Invalid request method'}, status=400)

    try:
        # ✅ Debugging Log
        logging.info(f"📌 Raw Request Body: {request.body}")

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            logging.error("❌ JSON Decode Error: Invalid JSON format")
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

        email = data.get('email')
        pdf_data = data.get('pdfData')

        # ✅ Debug: Log Extracted Values
        logging.info(f"✅ Extracted Email: {email}")
        logging.info(f"✅ PDF Data Received: {bool(pdf_data)}")

        if not email:
            logging.error("❌ Missing Email")
            return JsonResponse({'error': 'Missing email'}, status=400)

        if not pdf_data:
            logging.error("❌ Missing PDF Data")
            return JsonResponse({'error': 'Missing PDF data'}, status=400)

        # ✅ Convert Base64 to Binary PDF File
        pdf_bytes = base64.b64decode(pdf_data)
        pdf_filename = "summary_report.pdf"

        # ✅ Send Email via Azure Logic App
        post_data = {
            "recipients": email,
            "subject": f"Exposure Score Report - {datetime.now().strftime('%B %Y')}",
            "message_content": "Please find the attached exposure report.",
            "attachment_content": base64.b64encode(pdf_bytes).decode("utf-8"),
            "attachment_filename": pdf_filename,
            "function_name": "la-security-secops-email-endpoint-prod-northeurope",
            "verify_string": "2de87331-4c38-4c$$0-bbc2-d1b6dfb63d3b-#skbdhbf"
        }

        response = requests.post(LOGIC_APP_URL, json=post_data)
        response.raise_for_status()

        logging.info(f"✅ Email Sent Successfully! Status Code: {response.status_code}")
        return JsonResponse({'success': 'Email sent successfully!'}, status=200)

    except requests.exceptions.RequestException as e:
        logging.error(f"❌ Request error: {str(e)}")
        return JsonResponse({'error': f'Failed to send email: {str(e)}'}, status=500)

    except Exception as e:
        logging.error(f"❌ Unexpected error: {str(e)}")
        return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)


# @csrf_exempt
# def send_pdf_email(request):
#     if request.method != "POST":
#         return JsonResponse({'error': 'Invalid request method'}, status=400)

#     try:
#         # ✅ Log Raw Request Data for Debugging
#         logging.info(f"📌 Raw Request Body: {request.body}")

#         # Extract email and graphs data
#         data = json.loads(request.body)
#         email = data.get('email')
#         print(email,"email.......")
#         graphs_data = data.get('graphsData')
#         print(graphs_data, "graph dat......")


#         logging.info(f"✅ Extracted Email: {email}")
#         logging.info(f"✅ Extracted Graphs Count: {len(graphs_data) if graphs_data else 0}")

#         if not email or not graphs_data:
#             logging.error("❌ Missing email or graphs data")
#             return JsonResponse({'error': 'Missing email or graphs data'}, status=400)

#          # ✅ Debugging Graphs Data
#         for i, graph in enumerate(graphs_data):
#             logging.info(f"📊 Graph {i+1} Data (Base64 Length): {len(graph)}")


#         # Generate PDF from graphs
#         pdf_buffer = BytesIO()
#         pdf_canvas = canvas.Canvas(pdf_buffer, pagesize=letter)
#         width, height = letter

#         pdf_canvas.setFont("Helvetica-Bold", 16)
#         pdf_canvas.drawString(200, height - 50, f"Exposure Score Report - {datetime.now().strftime('%B %Y')}")

#         y_position = height - 100

#         for graph_data in graphs_data:
#             try:
#                 # Convert Base64 graph to an image
#                 graph_bytes = base64.b64decode(graph_data.split(',')[1])
#                 image = Image.open(BytesIO(graph_bytes))

#                 # Resize image for PDF
#                 image.thumbnail((500, 350))
#                 img_buffer = BytesIO()
#                 image.save(img_buffer, format="PNG")

#                 pdf_canvas.drawInlineImage(img_buffer, 50, y_position - 350, width=500, height=300)
#                 y_position -= 360

#                 if y_position < 100:  
#                     pdf_canvas.showPage()
#                     y_position = height - 100

#             except Exception as e:
#                 logging.error(f"❌ Error processing graph image: {str(e)}")

#         pdf_canvas.save()
#         pdf_content = pdf_buffer.getvalue()
#         pdf_buffer.close()

#         # Encode PDF to Base64
#         pdf_base64 = base64.b64encode(pdf_content).decode("utf-8")
#         pdf_filename = "summary_report.pdf"

#         # Send email via Azure Logic App
#         post_data = {
#             "recipients": email,
#             "subject": f"Exposure Score Report - {datetime.now().strftime('%B %Y')}",
#             "message_content": "Please find the attached exposure report.",
#             "attachment_content": pdf_base64,
#             "attachment_filename": pdf_filename,
#             "function_name": "la-security-secops-email-endpoint-prod-northeurope",
#             "verify_string": "2de87331-4c38-4c$$0-bbc2-d1b6dfb63d3b-#skbdhbf"
#         }

#         response = requests.post(LOGIC_APP_URL, json=post_data)
#         response.raise_for_status()

#         logging.info(f"✅ Email Sent Successfully! Status Code: {response.status_code}")
#         return JsonResponse({'success': 'Email sent successfully!'}, status=200)

#     except json.JSONDecodeError:
#         logging.error("❌ JSON Decode Error")
#         return JsonResponse({'error': 'Invalid JSON format'}, status=400)

#     except requests.exceptions.RequestException as e:
#         logging.error(f"❌ Request error: {str(e)}")
#         return JsonResponse({'error': f'Failed to send email: {str(e)}'}, status=500)

#     except Exception as e:
#         logging.error(f"❌ Unexpected error: {str(e)}")
#         return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)


def home(request):
    """Home page for file upload."""
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            file_path = uploaded_file.file.path
            
            # ✅ Encode file path properly
            encoded_file_path = urllib.parse.quote_plus(file_path)

            print(f"Redirecting to dashboard with file_path: {encoded_file_path}")  # Debugging

            try:
                # ✅ Correct Django URL reversal
                dashboard_url = reverse('dashboard', kwargs={'file_path': encoded_file_path})
                print(f"Resolved dashboard URL: {dashboard_url}")
                return redirect(dashboard_url)
            except Exception as e:
                print(f"Error resolving URL: {e}")

            # ✅ Fallback absolute URL for Azure
            azure_base_url = "https://as-antipiracy-phase1-dev-eceje4dqddb7gbev.centralindia-01.azurewebsites.net"
            return redirect(f"{azure_base_url}/dashboard/{encoded_file_path}/")

    else:
        form = UploadFileForm()

    return render(request, 'dashboard/home.html', {'form': form})


def read_excel_in_chunks(file_path, sheet_names=None, chunk_size=1000):
    """Read multiple Excel sheets in chunks using openpyxl."""
    
    # ✅ Load workbook and check available sheets
    workbook = load_workbook(file_path, read_only=True)
    print("Available Sheets in the Workbook:", workbook.sheetnames)

    # ✅ If sheet_names is None, read all sheets
    sheets_to_read = sheet_names if sheet_names else workbook.sheetnames

    for sheet in sheets_to_read:
        if sheet not in workbook.sheetnames:
            print(f"⚠️ Warning: Sheet '{sheet}' not found in the workbook. Skipping...")
            continue  # Skip missing sheets

        worksheet = workbook[sheet]

        # ✅ Ensure headers exist
        try:
            headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        except StopIteration:
            print(f"⚠️ Warning: Sheet '{sheet}' is empty. Skipping...")
            continue

        # ✅ Handle missing headers (replace None with generic names)
        headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(headers)]
        headers.append('SheetName')  # Add SheetName column to track source sheet
        
        chunk = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
            row_data = list(row) + [sheet]  # Append SheetName to row
            chunk.append(dict(zip(headers, row_data)))
            
            if i % chunk_size == 0:
                df_chunk = pd.DataFrame(chunk)
                
                
                if not df_chunk.empty and 'Identification Timestamp' in df_chunk.columns:
                    df_chunk['Identification Timestamp'] = pd.to_datetime(df_chunk['Identification Timestamp'], errors='coerce')

                yield df_chunk
                chunk = []

        if chunk:  # Yield remaining rows
            df_chunk = pd.DataFrame(chunk)
            
            if not df_chunk.empty and 'Identification Timestamp' in df_chunk.columns:
                df_chunk['Identification Timestamp'] = pd.to_datetime(df_chunk['Identification Timestamp'], errors='coerce')

            yield df_chunk

def generate_summary_graphs(summary_data):
    print("summary dataa..", summary_data)
    graphs = []
    # Graph 1: Sheet-wise Identification & Delisted
    if 'SheetName' in summary_data.columns and 'URL' in summary_data.columns:
        graphs.append(create_bar_chart(summary_data))

    # Graph 2: Monthly Identification & Removal Trends
    monthly_data = get_monthly_totals(summary_data)
    if not monthly_data.empty:
        graph2 = create_monthly_totals_line_plot(monthly_data)
        if graph2:
            graphs.append(graph2)

    # Graph 3: Social Media Platform Analysis
    social_media_data = get_social_media_platform_data(summary_data)
    if not social_media_data.empty:
        graph_social_media = create_social_media_platform_bar_chart(social_media_data)
        if graph_social_media:
            graphs.append(graph_social_media)

    #graph4
    matchday_summary = aggregate_matchday(summary_data)
    if not matchday_summary.empty:
        print("📊 Matchday Summary Ready for Graph Generation!", matchday_summary)  

        matchday_graph = create_matchday_line_plot(matchday_summary)

        if matchday_graph:
            print("✅ Matchday Graph Successfully Generated!")  
            graphs.append(matchday_graph)
        else:
            print("⚠️ Matchday Graph is Empty! Check Data or Rendering")
    else:
        print("⚠️ Matchday Summary is Empty! No Graph Generated.")



    # Graph 5: Top 5 Properties - Infringements and Removals
    if 'propertyname' in summary_data.columns and 'Status' in summary_data.columns and 'URL' in summary_data.columns:
        top_properties = summary_data.groupby('propertyname')['URL'].count().nlargest(5)
        removals = (
            summary_data[summary_data['Status'] == 'Approved']
            .groupby('propertyname')['URL']
            .count()
            .reindex(top_properties.index, fill_value=0)
        )
        fig, ax = plt.subplots(figsize=(10, 6), facecolor='#f8f9fa')
        x = np.arange(len(top_properties))
        bar_width = 0.3
        bar1 = ax.bar(x - bar_width / 2, top_properties, width=bar_width, label='Infringements', color='#ff6b6b', edgecolor='black', linewidth=0.7)
        bar2= ax.bar(x + bar_width / 2, removals, width=bar_width, label='Removals', color='#4a90e2', edgecolor='black', linewidth=0.7)

        ax.set_title('Top 5 Properties - Identification & Removals', fontsize=16)
        ax.set_xticks(x)

            # Adjust Y-axis limits to prevent overlap at top and bottom
        min_ylim = 1 if removals.min() > 0 else 0.5
        max_ylim = ax.get_ylim()[1] * 1.3  # Increased space above the highest bar
        ax.set_ylim(min_ylim, max_ylim)

         # Adding annotations
        for bar in bar1:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_properties), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#ff6b6b')
        for bar in bar2:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(removals), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#4a90e2')

        ax.set_xticklabels(top_properties.index, rotation=45, ha='right')
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        ax.set_facecolor('#f5f5f5')  # Light gray background inside the plot area
        ax.set_ylabel('Count')
        ax.legend()
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)

    # Graph 3: Top 5 Fixtures - Identification & Removals
    fixture_summary = summary_data.groupby(['fixtures', 'URL']).agg(
        removal_flag=('Status', lambda x: any(x.isin(['Approved', 'Removed'])))
    ).reset_index()

    fixture_summary = fixture_summary.groupby('fixtures').agg(
        total_urls=('URL', 'count'),
        removal_count=('removal_flag', 'sum')
    ).reset_index().sort_values(by='total_urls', ascending=False)

    top_fixtures = fixture_summary.nlargest(5, 'total_urls')[['fixtures', 'total_urls', 'removal_count']]
    
    if not top_fixtures.empty:
        fig, ax = plt.subplots(figsize=(12, 6), facecolor='#f8f9fa')
        x = np.arange(len(top_fixtures['fixtures']))
        bar_width = 0.3
         # Improved bar colors
        bar1 = ax.bar(x - bar_width / 2, top_fixtures['total_urls'], width=bar_width, label='Total URLs', color='#ff6b6b', edgecolor='black', linewidth=0.7)
        bar2 = ax.bar(x + bar_width / 2, top_fixtures['removal_count'], width=bar_width, label='Removal Count', color='#4a90e2', edgecolor='black', linewidth=0.7)

        ax.set_title("Top 5 Fixtures - Identification & Removal", fontsize=14, fontweight='bold')
        ax.set_ylabel('Count', fontsize=12, fontweight='bold')

        # Enhanced title and labels 
        ax.set_xticks(x)
        ax.set_xticklabels(top_fixtures['fixtures'], rotation=45, ha='right', fontsize=10)


        ax.grid(axis='y', linestyle='--', alpha=0.7)
        ax.set_facecolor('#f5f5f5')  # Light gray background inside the plot area
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        # Adding annotations
        for bar in bar1:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_fixtures['total_urls']), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#ff6b6b')
        for bar in bar2:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_fixtures['removal_count']), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#4a90e2')

        # Legend customization
        ax.legend(loc='upper right', fontsize=10, facecolor='white', edgecolor='black')

        plt.tight_layout()
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)


    #graph 4 sheet wise data
    
    return graphs



def generate_internet_graphs(summary_data, button):

    print("button",button)

    print("summary dataa..", summary_data)   
    graphs = []
    
    # Graph 1: Sheet-wise Identification & Delisted
    if 'SheetName' in summary_data.columns and 'URL' in summary_data.columns:
        graphs.append(create_bar_chart(summary_data))

    # Graph 2: Monthly Identification & Removal Trends
    monthly_data = get_monthly_totals(summary_data)
    if not monthly_data.empty:
        graph2 = create_monthly_totals_line_plot(monthly_data)
        if graph2:
            graphs.append(graph2)

    # Graph 3: Social Media Platform Analysis
    social_media_data = get_social_media_platform_data(summary_data)
    if not social_media_data.empty:
        graph_social_media = create_social_media_platform_bar_chart(social_media_data)
        if graph_social_media:
            graphs.append(graph_social_media)

    #gapgh4
    source_of_piracy = generate_top_10_piracy_sources_graph(summary_data)
    graphs.append(source_of_piracy)

    #graph5
    if button == "SocialMediaPlatforms":
        pass
    else:
        search_string = generate_top_10_search_treemap(summary_data)
        graphs.append(search_string)

    #graph4
    matchday_summary = aggregate_matchday(summary_data)
    if not matchday_summary.empty:
        print("📊 Matchday Summary Ready for Graph Generation!", matchday_summary)  

        matchday_graph = create_matchday_line_plot(matchday_summary)

        if matchday_graph:
            print("✅ Matchday Graph Successfully Generated!")  
            graphs.append(matchday_graph)
        else:
            print("⚠️ Matchday Graph is Empty! Check Data or Rendering")
    else:
        print("⚠️ Matchday Summary is Empty! No Graph Generated.")



    # Graph 5: Top 5 Properties - Infringements and Removals
    if 'propertyname' in summary_data.columns and 'Status' in summary_data.columns and 'URL' in summary_data.columns:
        top_properties = summary_data.groupby('propertyname')['URL'].count().nlargest(5)
        removals = (
            summary_data[summary_data['Status'] == 'Approved']
            .groupby('propertyname')['URL']
            .count()
            .reindex(top_properties.index, fill_value=0)
        )
        fig, ax = plt.subplots(figsize=(10, 6), facecolor='#f8f9fa')
        x = np.arange(len(top_properties))
        bar_width = 0.3
        bar1 = ax.bar(x - bar_width / 2, top_properties, width=bar_width, label='Infringements', color='#ff6b6b', edgecolor='black', linewidth=0.7)
        bar2= ax.bar(x + bar_width / 2, removals, width=bar_width, label='Removals', color='#4a90e2', edgecolor='black', linewidth=0.7)

        ax.set_title('Top 5 Properties - Identification & Removals', fontsize=16)
        ax.set_xticks(x)

            # Adjust Y-axis limits to prevent overlap at top and bottom
        min_ylim = 1 if removals.min() > 0 else 0.5
        max_ylim = ax.get_ylim()[1] * 1.3  # Increased space above the highest bar
        ax.set_ylim(min_ylim, max_ylim)

         # Adding annotations
        for bar in bar1:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_properties), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#ff6b6b')
        for bar in bar2:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(removals), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#4a90e2')

        ax.set_xticklabels(top_properties.index, rotation=45, ha='right')
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        ax.set_facecolor('#f5f5f5')  # Light gray background inside the plot area
        ax.set_ylabel('Count')
        ax.legend()
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)

    # Graph 3: Top 5 Fixtures - Identification & Removals
    fixture_summary = summary_data.groupby(['fixtures', 'URL']).agg(
        removal_flag=('Status', lambda x: any(x.isin(['Approved', 'Removed'])))
    ).reset_index()

    fixture_summary = fixture_summary.groupby('fixtures').agg(
        total_urls=('URL', 'count'),
        removal_count=('removal_flag', 'sum')
    ).reset_index().sort_values(by='total_urls', ascending=False)

    top_fixtures = fixture_summary.nlargest(5, 'total_urls')[['fixtures', 'total_urls', 'removal_count']]
    
    if not top_fixtures.empty:
        fig, ax = plt.subplots(figsize=(12, 6), facecolor='#f8f9fa')
        x = np.arange(len(top_fixtures['fixtures']))
        bar_width = 0.3
         # Improved bar colors
        bar1 = ax.bar(x - bar_width / 2, top_fixtures['total_urls'], width=bar_width, label='Total URLs', color='#ff6b6b', edgecolor='black', linewidth=0.7)
        bar2 = ax.bar(x + bar_width / 2, top_fixtures['removal_count'], width=bar_width, label='Removal Count', color='#4a90e2', edgecolor='black', linewidth=0.7)

        ax.set_title("Top 5 Fixtures - Identification & Removal", fontsize=14, fontweight='bold')
        ax.set_ylabel('Count', fontsize=12, fontweight='bold')

        # Enhanced title and labels 
        ax.set_xticks(x)
        ax.set_xticklabels(top_fixtures['fixtures'], rotation=45, ha='right', fontsize=10)


        ax.grid(axis='y', linestyle='--', alpha=0.7)
        ax.set_facecolor('#f5f5f5')  # Light gray background inside the plot area
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        # Adding annotations
        for bar in bar1:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_fixtures['total_urls']), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#ff6b6b')
        for bar in bar2:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02 * max(top_fixtures['removal_count']), 
                    f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=10, color='#4a90e2')

        # Legend customization
        ax.legend(loc='upper right', fontsize=10, facecolor='white', edgecolor='black')

        plt.tight_layout()
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)


    #graph 4 sheet wise data
    domain= generate_top_10_domains_bar_chart(summary_data)
    graphs.append(domain)
    
    return graphs


#  Function to generate Telegram graph summary
def generate_telegram_summary_graphs(summary_data):
    """Generate all Telegram-related summary graphs."""
    graphs = []
    print("coloumns listt---",summary_data.columns.tolist())
     # Graph 1: Monthly Identification & Removal Trends
    fixture_data = get_telegram_top_fixtures(summary_data)
    if not fixture_data.empty:
        graph2 =  create_telegram_top_fixtures_bar_chart(fixture_data)
        if graph2:
            graphs.append(graph2)
    #GRAPH2 : TREE GRAPH TOP10 
    top_10_channels =  create_treemap_chart_telegram(telegram_domains_by_subscribers(summary_data))
    graphs.append(top_10_channels)
    #graph3 : top_property_bar_chart
    top_property_bar_chart= create_top_property_bar_chart(get_top_telegram_properties(summary_data))
    graphs.append(top_property_bar_chart)
    #graph4 : 
    channel_type = create_channel_type_pie_chart(summary_data)
    graphs.append(channel_type)

    #graph 5
    Top_veiws =  top_fixtures_graph_donut_chart(top_fixtures_donut_chart(summary_data))
    graphs.append(Top_veiws)
    #graph6
    telegram_montly = telegram_monthly_totals_line_plot(telegram_monthly_totals(summary_data))
    graphs.append(telegram_montly)
    
    return graphs

def generate_mobileApp_graphs(summary_data):
    graphs = []
    #GRAPH1 : TREE GRAPH TOP10 
    top_5_MobileApp=  MobileApp_bar_chart(Mobile_properties_data(summary_data))
    graphs.append(top_5_MobileApp)

     # Graph 2: Monthly Identification & Removal Trends

    monthly_data = mobile_monthly_totals_line_plot(mobile_monthly_totals(summary_data))
    graphs.append(monthly_data)

    #graph3
    
    mobile_top5=  create_mobile_top_apps_bar_chart(Mobile_top_apps_by_downloads(summary_data))
    graphs.append(mobile_top5)


    #graph4
     #graph4
    mobile_matchday_summary = mobile_create_matchday_line_plot(mobile_aggregate_matchday(summary_data))
    graphs.append(mobile_matchday_summary)



    return graphs


def dashboard(request, file_path):
    """Dashboard view to display data."""
    decoded_file_path = urllib.parse.unquote(file_path)
    summary_data = pd.DataFrame()

    # ✅ Define correct titles for each section
    button_titles = {
        'Summary': 'Summary Dashboard',
        'internet': 'Internet Dashboard',
        'Mobile App':' Mobile App Dashbaord',
        'telegram': 'Telegram Dashboard',
        'SocialMediaPlatforms': 'Social Media Platforms Dashboard',
        'Premium': 'Premium Dashboard',
        'Send Mail': 'Email Dashboard'
    }
    buttons = list(button_titles.keys())  # Keep buttons dynamically from dict

    buttons = ['Summary', 'internet','Mobile App', 'telegram', 'SocialMediaPlatforms', 'Premium', 'Send Mail']

    button_filters = {
        'Summary': ['propertyname', 'fixtures'],
        'internet': ['propertyname', 'fixtures'],
        'Mobile App':['propertyname', 'fixtures'],
        'telegram': ['propertyname', 'fixtures'],
        'SocialMediaPlatforms': ['Platform', 'Engagement'],
        'Premium': ['Subscription Type', 'Duration']
    }
    
    selected_button = request.GET.get('button', 'Summary') 
    if selected_button == "Enforcement_Sheet_Infringing":  
        selected_button = "internet"
    elif selected_button == "MobileApp" :
        selected_button =  "MobileApplications"
    elif selected_button == "SocialMedia" :
        selected_button =  "SocialMediaPlatforms"

    
    applied_filters = {
        key: request.GET.get(key, '') for key in button_filters.get(selected_button, [])
    }
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    chunk_size = 1000
    graph_urls = []
    unique_values = {}
    widgets = {}

    workbook = load_workbook(decoded_file_path, read_only=True)
    sheet_names = workbook.sheetnames
    
    all_chunks = []
    # ✅ FIX: Ensure "Internet" reads from both `Enforcement_Sheet_Infringing` and `Enforcement_Sheet_Source`
    if selected_button == 'internet':
        print(selected_button, "selected button-----!")
        relevant_sheets = [sheet for sheet in ['Enforcement_Sheet_Infringing', 'Enforcement_Sheet_Source'] if sheet in sheet_names]
        print(relevant_sheets)
        if not relevant_sheets:
            print("⚠️ No valid sheets found for 'internet'. Skipping...")
            summary_data = pd.DataFrame()  # Return an empty DataFrame
        else:
            print(f"📂 Available Sheets for 'internet': {relevant_sheets}")

            for sheet in relevant_sheets:
                print(f"Reading Data from: {sheet}")
                for chunk in read_excel_in_chunks(decoded_file_path, sheet_names=[sheet], chunk_size=chunk_size):
                    if not chunk.empty:
                        # ✅ Extract unique values for filters
                        for column in button_filters.get(selected_button, []):
                            if column in chunk.columns:
                                if column not in unique_values:
                                    unique_values[column] = set()
                                unique_values[column].update(chunk[column].dropna().unique())

                        # ✅ Apply filters
                        for column, value in applied_filters.items():           
                            if value and column in chunk.columns:
                                chunk = chunk[chunk[column] == value]

                        # ✅ Apply date range filter
                        if start_date and end_date and 'Identification Timestamp' in chunk.columns:
                            chunk['Identification Timestamp'] = pd.to_datetime(chunk['Identification Timestamp'], errors='coerce')
                            chunk = chunk[(chunk['Identification Timestamp'] >= pd.to_datetime(start_date)) & 
                                          (chunk['Identification Timestamp'] <= pd.to_datetime(end_date))]

                        if not chunk.empty:
                            all_chunks.append(chunk)

            internet_data = pd.concat(all_chunks, ignore_index=True) if all_chunks else pd.DataFrame()

            unique_infringements = 0
            if 'Enforcement_Sheet_Infringing' in sheet_names:
                infringement_chunks = []
                for chunk in read_excel_in_chunks(decoded_file_path, sheet_names=['Enforcement_Sheet_Infringing'], chunk_size=chunk_size):
                    if not chunk.empty and 'URL' in chunk.columns:
                        infringement_chunks.append(chunk)

                if infringement_chunks:
                    infringement_data = pd.concat(infringement_chunks, ignore_index=True)
                    unique_infringements = infringement_data['URL'].nunique()

            widgets = {
                'Total Properties': internet_data['propertyname'].nunique() if 'propertyname' in internet_data.columns else "⚠️ Not Found",
                'Total Fixtures': internet_data['fixtures'].nunique() if 'fixtures' in internet_data.columns else "⚠️ Not Found",
                'No.Of Infringements': internet_data['URL'].nunique() if 'URL' in internet_data.columns else "⚠️ Not Found",
                'No.of Unique websites': internet_data['DomainName'].nunique() if 'DomainName' in internet_data.columns else "⚠️ Not Found",
                'No.Of unique infringing':unique_infringements ,
                '% Removal': round(
                    (internet_data.loc[internet_data['Status'].isin(['Approved', 'Removed']), 'URL'].nunique() /
                    internet_data['URL'].nunique() * 100) if 'Status' in internet_data.columns and internet_data['URL'].nunique() > 0 else 0, 3
                ),
                'No.Of Websites/Channels Suspended': internet_data.loc[internet_data['Status'].isin(['Approved', 'Removed']), 'DomainName'].nunique()
            }
            graph_urls.extend(generate_internet_graphs(internet_data, button= "internet"))

    else:
        # ✅ Default case for other sections (reading from a single sheet)
        for chunk in read_excel_in_chunks(decoded_file_path, sheet_names=[selected_button] if selected_button != 'Summary' else None, chunk_size=chunk_size):
            if not chunk.empty:
                
                # ✅ Extract unique values for filters
                for column in button_filters.get(selected_button, []):
                    if column in chunk.columns:
                        if column not in unique_values:
                            unique_values[column] = set()
                        unique_values[column].update(chunk[column].dropna().unique())

                # ✅ Apply filters
                for column, value in applied_filters.items():           
                    if value and column in chunk.columns:
                        chunk = chunk[chunk[column] == value]

                # ✅ Apply date range filter
                if start_date and end_date and 'Identification Timestamp' in chunk.columns:
                    chunk['Identification Timestamp'] = pd.to_datetime(chunk['Identification Timestamp'], errors='coerce')
                    chunk = chunk[(chunk['Identification Timestamp'] >= pd.to_datetime(start_date)) & 
                                  (chunk['Identification Timestamp'] <= pd.to_datetime(end_date))]

                if not chunk.empty:
                    all_chunks.append(chunk)

        summary_data = pd.concat(all_chunks, ignore_index=True) if all_chunks else pd.DataFrame()

    if not summary_data.empty:
        if selected_button == 'Summary':
            widgets = {
                'Total Properties': summary_data['propertyname'].nunique(),
                'Total Fixtures': summary_data['fixtures'].nunique(),
                'Total Infringements': summary_data['URL'].nunique(),
                'No.of Websites/channels': summary_data['DomainName'].nunique() ,
                '% Removal': round((summary_data.loc[summary_data['Status'].isin(['Approved', 'Removed']), 'URL'].nunique() /summary_data['URL'].nunique() * 100) if summary_data['URL'].nunique() > 0 else 0,3),
                'No.Of Websites/Channels/Suspended': summary_data.loc[summary_data['Status'].isin(['Approved', 'Removed']), 'DomainName'].nunique(),
            }
            graph_urls.extend(generate_summary_graphs(summary_data))

        elif selected_button == 'telegram':
            telegram_data = summary_data[summary_data['SheetName'] == 'telegram'] 

            print("get telegram data",telegram_data)
            telegram_data['views'] = pd.to_numeric(telegram_data['views'], errors='coerce')
            telegram_data['channelsubscribers'] = pd.to_numeric(telegram_data['channelsubscribers'], errors='coerce')
            approved_removed_filter = telegram_data[telegram_data['Status'].isin(['Approved', 'Removed'])]['URL'].nunique()
            total_infringements_telegram = telegram_data['URL'].nunique()
            print("---total_infringements_telegram",total_infringements_telegram)
            widgets = {
                'Total Properties': telegram_data['propertyname'].nunique(),
                'Total Fixtures': telegram_data['fixtures'].nunique(),
                'Total Infringements': telegram_data['URL'].nunique(),
                'Total Channels': telegram_data['DomainName'].nunique(),
                '% Removal': round((approved_removed_filter / total_infringements_telegram) * 100 if total_infringements_telegram > 0 else 0,3),
                'Total Views': telegram_data['views'].fillna(0).sum(),
                'Channels Suspended': telegram_data[telegram_data['ChannelStatus'].isin(['Suspended'])]['URL'].nunique(),
                'Total Subscribers': telegram_data['channelsubscribers'].fillna(0).sum(),
                'Impacted Subscribers': telegram_data[telegram_data['Status'].isin(['Approved', 'Removed'])]['channelsubscribers'].fillna(0).sum(),
            }
            graph_urls.extend(generate_telegram_summary_graphs(telegram_data))

        elif selected_button == 'MobileApplications':
            MobileApplications = summary_data[summary_data['SheetName'] == 'MobileApplications'] 
            if isinstance(MobileApplications, pd.DataFrame):
                if 'AppDownloads' in MobileApplications.columns:
                    MobileApplications['AppDownloads'] = pd.to_numeric(MobileApplications['AppDownloads'], errors='coerce')
                else:
                    print("Column 'AppDownloads' not found in the DataFrame")
            else:
                print("MobileApplications is not a DataFrame")
            MobileApplications_removed_filter = MobileApplications[MobileApplications['Status'].isin(['Approved', 'Removed'])]['URL'].nunique()
            total_infringements_MobileApp= MobileApplications['URL'].nunique()
            MobileApplications['AppDownloads'] = pd.to_numeric(MobileApplications['AppDownloads'], errors='coerce')

            widgets = {
                'No.of Platforms': MobileApplications['Platform'].nunique() if 'Platform' in MobileApplications.columns else 0,
                'No.Of Mobile Apps': MobileApplications['URL'].nunique() if 'URL' in MobileApplications.columns else 0,
                'No.Of Unique Domains': MobileApplications['DomainName'].nunique() if 'DomainName' in MobileApplications.columns else 0,
                '% Removals': round((MobileApplications_removed_filter /total_infringements_MobileApp) * 100 if total_infringements_MobileApp > 0 else 0,3),
                'No.of Downloads': MobileApplications['AppDownloads'].fillna(0).sum(),
            }
            graph_urls.extend(generate_mobileApp_graphs(MobileApplications))


        elif selected_button == 'SocialMediaPlatforms':
            Social_Media_platform = summary_data[summary_data['SheetName'] == 'SocialMediaPlatforms'] 

            Social_Media_platform['views'] = pd.to_numeric(Social_Media_platform['views'], errors='coerce')
            approved_removed_filter = Social_Media_platform[Social_Media_platform['Status'].isin(['Approved', 'Removed'])]['URL'].nunique()
            total_infringements_telegram = Social_Media_platform['URL'].nunique()
            Social_Media_platform['channelsubscribers'] = pd.to_numeric(Social_Media_platform['channelsubscribers'], errors='coerce')

            widgets = {
                'Total Properties': Social_Media_platform['propertyname'].nunique() if 'propertyname' in Social_Media_platform.columns else "⚠️ Not Found",
                'Total Fixtures': Social_Media_platform['fixtures'].nunique() if 'fixtures' in Social_Media_platform.columns else "⚠️ Not Found",
                'No.Of Infringements': Social_Media_platform['URL'].nunique() if 'URL' in Social_Media_platform.columns else "⚠️ Not Found",
                'No.of Unique websites': Social_Media_platform['DomainName'].nunique() if 'DomainName' in Social_Media_platform.columns else "⚠️ Not Found",
                'Total Views': Social_Media_platform['views'].fillna(0).sum(),
                '% Removal': round(
                    (Social_Media_platform.loc[Social_Media_platform['Status'].isin(['Approved', 'Removed']), 'URL'].nunique() /
                    Social_Media_platform['URL'].nunique() * 100) if 'Status' in Social_Media_platform.columns and Social_Media_platform['URL'].nunique() > 0 else 0, 3
                ),
                'No.Of impacted subscribers': Social_Media_platform['channelsubscribers'].fillna(0).sum(),
                'No.Of Websites/Channels Suspended': Social_Media_platform.loc[Social_Media_platform['Status'].isin(['Approved', 'Removed']), 'DomainName'].nunique()
            }

            graph_urls.extend(generate_internet_graphs(Social_Media_platform, button= "SocialMediaPlatforms"))
            

    paginator = Paginator(summary_data.values.tolist(), 200)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    paginated_data = pd.DataFrame(page_obj.object_list, columns=summary_data.columns) if not summary_data.empty else None

    return render(request, 'dashboard/dashboard.html', {
        'buttons': buttons,
        'selected_button': selected_button,
        'filters': button_filters.get(selected_button, []),
        'applied_filters': applied_filters,
        'unique_values': unique_values,
        'paginated_data': paginated_data,
        'page_obj': page_obj,
        'widgets': widgets,
        'graphs': graph_urls,
        'start_date': start_date,
        'end_date': end_date,
    })
