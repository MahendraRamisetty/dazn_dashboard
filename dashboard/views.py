from django.shortcuts import render, redirect
from .forms import UploadFileForm
from .models import UploadedFile
import pandas as pd
import urllib.parse
from openpyxl import load_workbook
from django.core.paginator import Paginator
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for Matplotlib
import matplotlib.pyplot as plt
import io
import base64


def home(request):
    """Home page for file upload."""
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            file_path = uploaded_file.file.path
            encoded_file_path = urllib.parse.quote(file_path)
            return redirect('dashboard', file_path=encoded_file_path)
    else:
        form = UploadFileForm()
    return render(request, 'dashboard/home.html', {'form': form})


def read_excel_in_chunks(file_path, sheet_name=None, chunk_size=1000):
    """Read an Excel sheet in chunks using openpyxl."""
    workbook = load_workbook(file_path, read_only=True)
    sheets = workbook.sheetnames if sheet_name is None else [sheet_name]
    for sheet in sheets:
        worksheet = workbook[sheet]
        headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
        chunk = []
        for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
            chunk.append(dict(zip(headers, row)))
            if i % chunk_size == 0:
                yield pd.DataFrame(chunk)
                chunk = []
        if chunk:
            yield pd.DataFrame(chunk)


def generate_graph(data, column_name, graph_type):
    """Generate graphs dynamically based on the data."""
    fig, ax = plt.subplots(figsize=(8, 4))
    img = io.BytesIO()

    if column_name in data.columns:
        if graph_type == 'bar':
            data_counts = data[column_name].value_counts()
            data_counts.plot(kind='bar', ax=ax, color="skyblue")
            ax.set_title(f"Bar Chart of {column_name}")
        elif graph_type == 'pie':
            data_counts = data[column_name].value_counts()
            data_counts.plot(kind='pie', ax=ax, autopct='%1.1f%%', startangle=90)
            ax.set_title(f"Pie Chart of {column_name}")
        elif graph_type == 'line':
            data[column_name].plot(ax=ax)
            ax.set_title(f"Line Chart of {column_name}")

    plt.tight_layout()
    plt.savefig(img, format='png')
    img.seek(0)
    graph_url = base64.b64encode(img.getvalue()).decode()
    plt.close(fig)
    return f"data:image/png;base64,{graph_url}"


import numpy as np


def generate_summary_graphs(summary_data):
    graphs = []

    # Graph 1: Top 5 Properties - Infringements and Removals
    if 'propertyname' in summary_data.columns and 'Status' in summary_data.columns and 'URL' in summary_data.columns:
        top_properties = summary_data.groupby('propertyname')['URL'].count().nlargest(5)
        removals = (
            summary_data[summary_data['Status'] == 'Approved']
            .groupby('propertyname')['URL']
            .count()
            .reindex(top_properties.index, fill_value=0)
        )
        fig, ax = plt.subplots(figsize=(10, 6))
        x = np.arange(len(top_properties))
        ax.bar(x - 0.2, top_properties, width=0.4, label='Infringements', color='blue')
        ax.bar(x + 0.2, removals, width=0.4, label='Removals', color='orange')
        ax.set_title('Top 5 Properties - Identification & Removals', fontsize=16)
        ax.set_xticks(x)
        ax.set_xticklabels(top_properties.index, rotation=45, ha='right')
        ax.set_ylabel('Count')
        ax.legend()
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)

    # Graph 2: Status Distribution (Donut Chart)
    if 'Status' in summary_data.columns:
        fig, ax = plt.subplots(figsize=(8, 8))
        status_counts = summary_data['Status'].value_counts()
        wedges, texts, autotexts = ax.pie(
            status_counts,
            labels=status_counts.index,
            autopct='%1.1f%%',
            startangle=140,
            colors=plt.cm.Paired.colors
        )
        # Donut effect
        center_circle = plt.Circle((0, 0), 0.70, fc='white')
        fig.gca().add_artist(center_circle)
        ax.set_title('Status Distribution', fontsize=16)
        img = io.BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        img.seek(0)
        graphs.append(base64.b64encode(img.getvalue()).decode())
        plt.close(fig)

    return graphs



def dashboard(request, file_path):
    """Dashboard view to display data."""
    decoded_file_path = urllib.parse.unquote(file_path)

    buttons = ['Summary', 'Enforcement_Sheet_Infringing', 'Enforcement_Sheet_Source', 
               'Telegram', 'SocialMediaPlatforms', 'Premium']

    button_filters = {
        'Summary': ['propertyname', 'fixtures',],
        'Enforcement_Sheet_Infringing': ['propertyname', 'fixtures'],
        'Enforcement_Sheet_Source': ['Source Type', 'Date'],
        'Telegram': ['Message Type', 'User ID'],
        'SocialMediaPlatforms': ['Platform', 'Engagement'],
        'Premium': ['Subscription Type', 'Duration']
    }

    selected_button = request.GET.get('button', 'Summary')
    applied_filters = {
        key: request.GET.get(key, '') for key in button_filters.get(selected_button, [])
    }

    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    chunk_size = 1000
    graph_urls = []
    unique_values = {}
    widgets = {}

    workbook = load_workbook(decoded_file_path, read_only=True)
    sheet_names = workbook.sheetnames

    all_chunks = []
    for chunk in read_excel_in_chunks(decoded_file_path, sheet_name=selected_button if selected_button != 'Summary' else None, chunk_size=chunk_size):
        if not chunk.empty:
            for column in button_filters.get(selected_button, []):
                if column in chunk.columns:
                    if column not in unique_values:
                        unique_values[column] = set()
                    unique_values[column].update(chunk[column].dropna().unique())

            # Apply filters
            for column, value in applied_filters.items():
                if value and column in chunk.columns:
                    chunk = chunk[chunk[column] == value]

            # Apply date range filter
            if selected_button == 'Summary' and start_date and end_date:
                if 'Identification Timestamp' in chunk.columns:
                    chunk['Identification Timestamp'] = pd.to_datetime(chunk['Identification Timestamp'], errors='coerce')
                    chunk = chunk[(chunk['Identification Timestamp'] >= pd.to_datetime(start_date)) & 
                                  (chunk['Identification Timestamp'] <= pd.to_datetime(end_date))]

            if not chunk.empty:
                all_chunks.append(chunk)

    unique_values = {key: list(value) for key, value in unique_values.items()}
    summary_data = pd.concat(all_chunks, ignore_index=True) if all_chunks else pd.DataFrame()

    if not summary_data.empty:
        if selected_button == 'Summary':
            widgets = {
                'Total Properties': summary_data['propertyname'].nunique(),
                'Total Fixtures': summary_data['fixtures'].nunique(),
                'Total Infringements': summary_data['URL'].nunique(),
                'Removal Percentage': (summary_data.loc[summary_data['Status'].isin(['Approved', 'Removed']), 'URL'].nunique() /
                                       summary_data['URL'].nunique() * 100) if summary_data['URL'].nunique() > 0 else 0,
            }
            graph_urls.extend(generate_summary_graphs(summary_data))
        elif selected_button == 'Enforcement_Sheet_Infringing':
            widgets = {
                'Total URLs': summary_data['URL'].nunique() if 'URL' in summary_data.columns else 0,
                'Total Domains': summary_data['Domain Name'].nunique() if 'Domain Name' in summary_data.columns else 0,
            }
            graph_urls.append(generate_graph(summary_data, 'URL', 'bar'))
            graph_urls.append(generate_graph(summary_data, 'Domain Name', 'line'))
        elif selected_button == 'Enforcement_Sheet_Source':
            widgets = {
                'Total Sources': summary_data['Source Type'].nunique() if 'Source Type' in summary_data.columns else 0,
                'Earliest Date': summary_data['Date'].min() if 'Date' in summary_data.columns else 'N/A',
                'Latest Date': summary_data['Date'].max() if 'Date' in summary_data.columns else 'N/A',
            }
            graph_urls.append(generate_graph(summary_data, 'Source Type', 'bar'))
            graph_urls.append(generate_graph(summary_data, 'Date', 'line'))
        elif selected_button == 'Telegram':
            widgets = {
                'Message Types': summary_data['Message Type'].nunique() if 'Message Type' in summary_data.columns else 0,
                'Active Users': summary_data['User ID'].nunique() if 'User ID' in summary_data.columns else 0,
            }
            graph_urls.append(generate_graph(summary_data, 'Message Type', 'bar'))
            graph_urls.append(generate_graph(summary_data, 'User ID', 'line'))
        elif selected_button == 'Premium':
            widgets = {
                'Subscription Types': summary_data['Subscription Type'].nunique() if 'Subscription Type' in summary_data.columns else 0,
                'Average Duration': summary_data['Duration'].mean() if 'Duration' in summary_data.columns else 0,
            }
            graph_urls.append(generate_graph(summary_data, 'Subscription Type', 'pie'))
            graph_urls.append(generate_graph(summary_data, 'Duration', 'bar'))

    paginator = Paginator(summary_data.values.tolist(), 200)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    paginated_data = pd.DataFrame(page_obj.object_list, columns=summary_data.columns) if not summary_data.empty else None

    return render(request, 'dashboard/dashboard.html', {
        'buttons': buttons,
        'selected_button': selected_button,
        'filters': button_filters.get(selected_button, []),
        'applied_filters': applied_filters,
        'unique_values': unique_values,
        'paginated_data': paginated_data,
        'page_obj': page_obj,
        'widgets': widgets,
        'graphs': graph_urls,
        'start_date': start_date,
        'end_date': end_date,
    })
